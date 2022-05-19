from xlwt import Workbook
from openpyxl import load_workbook, Workbook


def save_result_to_file(result, iteration, sheet_name):
    try:
        wb = load_workbook('./results.xlsx')
        sheet = wb.get_sheet_by_name(sheet_name)
    except:
        wb = Workbook()
        sheet = wb.active
        sheet.title = sheet_name

    cell = sheet.cell(row=iteration, column=1)
    cell.value = result

    wb.save('./results.xlsx')
